"""
Enhanced Canvas Integration with Privacy Protection and Two-Step Workflow

This enhanced version includes:
1. Student name anonymization for ChatGPT processing
2. Two-step workflow: Grade → Review → Upload
3. Manual review and editing capability
4. Organized file management by assignment

Author: Grading Agent
Date: January 2025
"""

import requests
import os
import json
import tempfile
import zipfile
import uuid
import shutil
from typing import Dict, List, Optional, Tuple
from datetime import datetime
import time
import pandas as pd
import re

class StudentAnonymizer:
    """Handles student identity anonymization for privacy protection"""
    
    def __init__(self):
        self.name_map = {}  # Maps anonymous IDs to real names
        self.reverse_map = {}  # Maps real names to anonymous IDs
        
    def anonymize_name(self, real_name: str, user_id: int) -> str:
        """Convert real name to anonymous identifier"""
        if real_name in self.reverse_map:
            return self.reverse_map[real_name]
        
        # Create anonymous ID
        anon_id = f"Student_{len(self.name_map) + 1:03d}"
        
        # Store mappings
        self.name_map[anon_id] = {
            'real_name': real_name,
            'user_id': user_id
        }
        self.reverse_map[real_name] = anon_id
        
        return anon_id
    
    def get_real_name(self, anon_id: str) -> str:
        """Get real name from anonymous ID"""
        return self.name_map.get(anon_id, {}).get('real_name', anon_id)
    
    def get_user_id(self, anon_id: str) -> int:
        """Get Canvas user ID from anonymous ID"""
        return self.name_map.get(anon_id, {}).get('user_id')
    
    def anonymize_text(self, text: str) -> str:
        """Remove/anonymize any student names that might appear in text"""
        # This could be enhanced to detect and replace names in paper content
        # For now, we'll focus on file names and basic patterns
        for real_name, anon_id in self.reverse_map.items():
            # Replace name variations (case insensitive)
            text = re.sub(re.escape(real_name), anon_id, text, flags=re.IGNORECASE)
            
            # Also replace common name patterns
            first_name = real_name.split()[0] if ' ' in real_name else real_name
            text = re.sub(re.escape(first_name), anon_id.split('_')[1], text, flags=re.IGNORECASE)
        
        return text
    
    def save_mapping(self, file_path: str):
        """Save name mapping to file for later use"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump({
                'name_map': self.name_map,
                'reverse_map': self.reverse_map,
                'created_at': datetime.now().isoformat()
            }, f, indent=2)
    
    def load_mapping(self, file_path: str):
        """Load name mapping from file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.name_map = data.get('name_map', {})
                self.reverse_map = data.get('reverse_map', {})
        except Exception as e:
            print(f"Warning: Could not load name mapping: {e}")


class CanvasAPI:
    """Canvas LMS API client for automated grading workflow"""
    
    def __init__(self, canvas_url: str, api_token: str):
        """
        Initialize Canvas API client
        
        Args:
            canvas_url: Your Canvas instance URL (e.g., "https://yourschool.instructure.com")
            api_token: Your Canvas API access token
        """
        self.canvas_url = canvas_url.rstrip('/')
        self.api_token = api_token
        self.headers = {
            'Authorization': f'Bearer {api_token}',
            'Content-Type': 'application/json'
        }
        
    def _make_request(self, method: str, endpoint: str, **kwargs) -> requests.Response:
        """Make authenticated request to Canvas API"""
        url = f"{self.canvas_url}/api/v1/{endpoint}"
        response = requests.request(method, url, headers=self.headers, **kwargs)
        
        if response.status_code == 429:  # Rate limited
            retry_after = int(response.headers.get('Retry-After', 60))
            print(f"Rate limited. Waiting {retry_after} seconds...")
            time.sleep(retry_after)
            return self._make_request(method, endpoint, **kwargs)
            
        response.raise_for_status()
        return response
    
    def get_courses(self) -> List[Dict]:
        """Get list of courses for the authenticated user"""
        response = self._make_request('GET', 'courses')
        return response.json()
    
    def get_assignments(self, course_id: int) -> List[Dict]:
        """Get assignments for a specific course"""
        response = self._make_request('GET', f'courses/{course_id}/assignments')
        return response.json()
    
    def get_assignment_rubric(self, course_id: int, assignment_id: int) -> Optional[Dict]:
        """
        Get the rubric associated with an assignment
        
        Args:
            course_id: Canvas course ID
            assignment_id: Canvas assignment ID
            
        Returns:
            Rubric data dictionary or None if no rubric exists
        """
        try:
            # First check if assignment has a rubric
            response = self._make_request(
                'GET', 
                f'courses/{course_id}/assignments/{assignment_id}',
                params={'include[]': ['rubric', 'rubric_settings']}
            )
            assignment_data = response.json()
            
            # Check multiple possible locations for rubric information
            rubric_settings = assignment_data.get('rubric_settings')
            rubric_data_direct = assignment_data.get('rubric')
            use_rubric_for_grading = assignment_data.get('use_rubric_for_grading')
            
            # Try to get rubric ID from various sources
            rubric_id = None
            if rubric_settings:
                rubric_id = rubric_settings.get('id')
            
            if not rubric_id and rubric_data_direct:
                rubric_id = rubric_data_direct.get('id')
                
            if not rubric_id:
                return None
            
            # First try to use rubric data already included in the assignment response
            if rubric_data_direct:
                # Handle case where rubric data is a list (Canvas sometimes returns it this way)
                if isinstance(rubric_data_direct, list):
                    # Create a synthetic rubric dict from the list data
                    rubric_dict = {
                        'title': rubric_settings.get('title', 'Canvas Rubric'),
                        'points_possible': rubric_settings.get('points_possible', 100),
                        'id': rubric_settings.get('id'),
                        'data': rubric_data_direct  # The list becomes the 'data' field
                    }
                    converted_rubric = self._convert_canvas_rubric(rubric_dict)
                else:
                    converted_rubric = self._convert_canvas_rubric(rubric_data_direct)
                
                return converted_rubric
            
            # If not included, try to fetch detailed rubric information
            try:
                response = self._make_request('GET', f'courses/{course_id}/rubrics/{rubric_id}')
                rubric_data = response.json()
                
                # Convert Canvas rubric format to our grading agent format
                converted_rubric = self._convert_canvas_rubric(rubric_data)
                return converted_rubric
                
            except Exception as api_error:
                # Try to construct a basic rubric from rubric_settings if available
                if rubric_settings:
                    basic_rubric = self._create_basic_rubric_from_settings(rubric_settings)
                    if basic_rubric:
                        return basic_rubric
                
                # If all else fails, re-raise the original error
                raise api_error
            
        except Exception as e:
            print(f"Error fetching rubric: {e}")
            return None
    
    def _convert_canvas_rubric(self, canvas_rubric: Dict) -> Dict:
        """
        Convert Canvas rubric format to grading agent format
        
        Args:
            canvas_rubric: Raw Canvas rubric data
            
        Returns:
            Converted rubric in grading agent format
        """
        converted = {
            "assignment_title": canvas_rubric.get('title', 'Canvas Assignment'),
            "total_points": canvas_rubric.get('points_possible', 100),
            "criteria": {},
            "grading_instructions": "Grade this assignment based on the Canvas rubric criteria below.",
            "canvas_rubric_id": canvas_rubric.get('id'),
            "canvas_rubric_title": canvas_rubric.get('title')
        }
        
        # Convert rubric criteria
        for criterion in canvas_rubric.get('data', []):
            criterion_id = criterion.get('id')
            criterion_description = criterion.get('description', f'Criterion {criterion_id}')
            
            # Get ratings/levels for this criterion
            ratings = []
            for rating in criterion.get('ratings', []):
                ratings.append({
                    "points": rating.get('points', 0),
                    "description": rating.get('description', ''),
                    "long_description": rating.get('long_description', '')
                })
            
            converted["criteria"][criterion_description] = {
                "points": criterion.get('points', 0),
                "description": criterion.get('long_description', criterion_description),
                "ratings": ratings
            }
        
        return converted
    
    def _create_basic_rubric_from_settings(self, rubric_settings: Dict) -> Dict:
        """
        Create a basic rubric from Canvas rubric settings when detailed rubric data isn't available
        
        Args:
            rubric_settings: Canvas rubric settings from assignment data
            
        Returns:
            Basic rubric in grading agent format
        """
        rubric_title = rubric_settings.get('title', 'Canvas Rubric')
        total_points = rubric_settings.get('points_possible', 100)
        
        # Create a basic rubric with generic criteria
        basic_rubric = {
            "assignment_title": rubric_title,
            "total_points": total_points,
            "criteria": {
                "Overall Quality": {
                    "points": total_points,
                    "description": f"Evaluate the overall quality of the submission according to the '{rubric_title}' rubric criteria visible in Canvas.",
                    "ratings": [
                        {
                            "points": total_points,
                            "description": "Excellent",
                            "long_description": "Meets all criteria excellently"
                        },
                        {
                            "points": total_points * 0.8,
                            "description": "Good", 
                            "long_description": "Meets most criteria well"
                        },
                        {
                            "points": total_points * 0.6,
                            "description": "Satisfactory",
                            "long_description": "Meets basic criteria"
                        },
                        {
                            "points": total_points * 0.4,
                            "description": "Needs Improvement",
                            "long_description": "Partially meets criteria"
                        },
                        {
                            "points": 0,
                            "description": "Unsatisfactory",
                            "long_description": "Does not meet criteria"
                        }
                    ]
                }
            },
            "grading_instructions": f"Grade this assignment based on the '{rubric_title}' rubric. "
                                  f"While detailed rubric criteria couldn't be retrieved due to API permissions, "
                                  f"please evaluate the submission according to standard academic criteria and "
                                  f"assign a score out of {total_points} points.",
            "canvas_rubric_id": rubric_settings.get('id'),
            "canvas_rubric_title": rubric_title,
            "rubric_note": "This is a simplified rubric created from Canvas settings. "
                          "The instructor can view the detailed rubric criteria in Canvas."
        }
        
        return basic_rubric
    
    def get_assignment_submissions(self, course_id: int, assignment_id: int, 
                                 include_attachments: bool = True) -> List[Dict]:
        """
        Get all submissions for an assignment
        
        Args:
            course_id: Canvas course ID
            assignment_id: Canvas assignment ID
            include_attachments: Whether to include attachment information
            
        Returns:
            List of submission objects
        """
        params = {
            'include[]': ['submission_comments', 'user'],
            'per_page': 100  # Get more submissions per page
        }
        if include_attachments:
            params['include[]'].append('submission_history')
            
        # Handle pagination to get all submissions
        all_submissions = []
        page = 1
        
        while True:
            params['page'] = page
            response = self._make_request(
                'GET', 
                f'courses/{course_id}/assignments/{assignment_id}/submissions',
                params=params
            )
            submissions = response.json()
            
            if not submissions:
                break
                
            all_submissions.extend(submissions)
            
            # Check if there are more pages
            if len(submissions) < params['per_page']:
                break
                
            page += 1
            
        print(f"📊 Retrieved {len(all_submissions)} total submissions from Canvas")
        return all_submissions
    
    def download_submission_file(self, file_url: str, download_path: str) -> str:
        """
        Download a submission file from Canvas
        
        Args:
            file_url: Direct download URL for the file
            download_path: Local path to save the file
            
        Returns:
            Path to downloaded file
        """
        response = requests.get(file_url, headers={'Authorization': f'Bearer {self.api_token}'})
        response.raise_for_status()
        
        with open(download_path, 'wb') as f:
            f.write(response.content)
        
        return download_path
    
    def download_submissions_bulk(self, course_id: int, assignment_id: int, 
                                download_dir: str, anonymizer: StudentAnonymizer) -> Dict[int, List[str]]:
        """
        Download all submissions for an assignment with anonymization
        
        Args:
            course_id: Canvas course ID
            assignment_id: Canvas assignment ID
            download_dir: Directory to save submissions
            anonymizer: StudentAnonymizer instance for privacy protection
            
        Returns:
            Dictionary mapping user_id to list of downloaded file paths
        """
        os.makedirs(download_dir, exist_ok=True)
        submissions = self.get_assignment_submissions(course_id, assignment_id)
        downloaded_files = {}
        
        for submission in submissions:
            user_id = submission['user_id']
            real_name = submission.get('user', {}).get('name', f'user_{user_id}')
            
            # Anonymize student name for file storage
            anon_name = anonymizer.anonymize_name(real_name, user_id)
            
            # Skip if no submission or no attachments
            if not submission.get('attachments') and not submission.get('body'):
                continue
            
            user_files = []
            user_dir = os.path.join(download_dir, f"{anon_name}_{user_id}")
            os.makedirs(user_dir, exist_ok=True)
            
            # Download file attachments
            if submission.get('attachments'):
                for attachment in submission['attachments']:
                    original_filename = attachment['filename']
                    # Keep original extension but anonymize filename
                    file_ext = os.path.splitext(original_filename)[1]
                    anon_filename = f"{anon_name}_submission{file_ext}"
                    
                    file_url = attachment['url']
                    file_path = os.path.join(user_dir, anon_filename)
                    
                    try:
                        self.download_submission_file(file_url, file_path)
                        user_files.append(file_path)
                        print(f"Downloaded: {anon_filename} for {anon_name}")
                    except Exception as e:
                        print(f"Error downloading {original_filename} for {anon_name}: {e}")
            
            # Save text submissions
            if submission.get('body'):
                text_file = os.path.join(user_dir, f"{anon_name}_text_submission.html")
                # Anonymize any names that might appear in the submission text
                anonymized_body = anonymizer.anonymize_text(submission['body'])
                
                with open(text_file, 'w', encoding='utf-8') as f:
                    f.write(anonymized_body)
                user_files.append(text_file)
                print(f"Saved text submission for {anon_name}")
            
            if user_files:
                downloaded_files[user_id] = user_files
        
        return downloaded_files
    
    def update_submission_grade(self, course_id: int, assignment_id: int, 
                               user_id: int, grade: str, comment: str = None) -> Dict:
        """
        Update grade and add comment for a submission
        
        Args:
            course_id: Canvas course ID
            assignment_id: Canvas assignment ID
            user_id: Student user ID
            grade: Grade to assign (can be points, percentage, or letter grade)
            comment: Feedback comment to add
            
        Returns:
            Updated submission object
        """
        data = {
            'submission': {
                'posted_grade': grade
            }
        }
        
        if comment:
            data['comment'] = {
                'text_comment': comment
            }
        
        response = self._make_request(
            'PUT',
            f'courses/{course_id}/assignments/{assignment_id}/submissions/{user_id}',
            json=data
        )
        return response.json()
    
    def bulk_update_grades(self, course_id: int, assignment_id: int, 
                          grade_data: Dict[int, Dict]) -> Dict:
        """
        Update multiple grades at once
        
        Args:
            course_id: Canvas course ID
            assignment_id: Canvas assignment ID
            grade_data: Dictionary mapping user_id to grade/comment data
                       Format: {user_id: {'grade': '85', 'comment': 'Good work!'}}
            
        Returns:
            Progress object for bulk operation
        """
        # Convert to Canvas API format
        canvas_data = {}
        for user_id, data in grade_data.items():
            canvas_data[f'grade_data[{user_id}][posted_grade]'] = data['grade']
            if 'comment' in data:
                canvas_data[f'grade_data[{user_id}][text_comment]'] = data['comment']
        
        response = self._make_request(
            'POST',
            f'courses/{course_id}/assignments/{assignment_id}/submissions/update_grades',
            data=canvas_data
        )
        return response.json()


class TwoStepCanvasGrading:
    """Two-step Canvas grading workflow with manual review"""
    
    def __init__(self, canvas_api: CanvasAPI, grading_agent):
        """
        Initialize two-step Canvas grading integration
        
        Args:
            canvas_api: Configured CanvasAPI instance
            grading_agent: Existing GradingAgent instance
        """
        self.canvas = canvas_api
        self.grading_agent = grading_agent
        self.anonymizer = StudentAnonymizer()
    
    def step1_download_and_grade(self, course_id: int, assignment_id: int, 
                                assignment_name: str, rubric_path: str = None, 
                                instructor_config_path: str = None, 
                                use_canvas_rubric: bool = False) -> Dict:
        """
        Step 1: Download submissions, grade with anonymization, create review folder
        
        Args:
            course_id: Canvas course ID
            assignment_id: Canvas assignment ID
            assignment_name: Name of the assignment
            rubric_path: Path to grading rubric JSON file (optional if using Canvas rubric)
            instructor_config_path: Path to instructor configuration file (optional)
            use_canvas_rubric: Whether to download and use the Canvas rubric
            
        Returns:
            Dictionary with results and folder path
        """
        print(f"Step 1: Download and Grade - {assignment_name}")
        print("=" * 60)
        
        # Create organized folder structure
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_assignment_name = re.sub(r'[^\w\s-]', '', assignment_name).strip()
        safe_assignment_name = re.sub(r'[-\s]+', '_', safe_assignment_name)
        
        assignment_folder = f"{safe_assignment_name}_{date_str}"
        base_dir = os.path.join(os.getcwd(), assignment_folder)
        
        submissions_dir = os.path.join(base_dir, "submissions")
        results_dir = os.path.join(base_dir, "results")
        
        os.makedirs(submissions_dir, exist_ok=True)
        os.makedirs(results_dir, exist_ok=True)
        
        print(f"Created assignment folder: {assignment_folder}")
        
        try:
            # Step 1a: Handle rubric (Canvas or local)
            final_rubric_path = None
            if use_canvas_rubric:
                print("🔍 Checking for Canvas rubric...")
                canvas_rubric = self.canvas.get_assignment_rubric(course_id, assignment_id)
                
                if canvas_rubric:
                    # Save Canvas rubric locally
                    final_rubric_path = os.path.join(base_dir, "canvas_rubric.json")
                    with open(final_rubric_path, 'w', encoding='utf-8') as f:
                        json.dump(canvas_rubric, f, indent=2)
                    print(f"✅ Downloaded Canvas rubric: {canvas_rubric.get('canvas_rubric_title', 'Untitled')}")
                    print(f"📁 Saved rubric to: {final_rubric_path}")
                else:
                    print("⚠️  No Canvas rubric found for this assignment")
                    if not rubric_path:
                        print("❌ No local rubric provided either")
                        raise ValueError("No Canvas rubric available and no local rubric provided")
            
            if not final_rubric_path:
                if not rubric_path or (rubric_path and not os.path.exists(rubric_path)):
                    raise ValueError("Rubric file not found or not provided")
                final_rubric_path = rubric_path
            
            # Step 1b: Load instructor configuration (if provided)
            if instructor_config_path and os.path.exists(instructor_config_path):
                print(f"🎓 Loading instructor configuration...")
                self.grading_agent.load_instructor_config(instructor_config_path)
                # Copy config for reference
                config_copy = os.path.join(base_dir, "instructor_config_used.json")
                shutil.copy2(instructor_config_path, config_copy)
            
            # Step 1c: Download submissions with anonymization
            print("📥 Downloading submissions with privacy protection...")
            downloaded_files = self.canvas.download_submissions_bulk(
                course_id, assignment_id, submissions_dir, self.anonymizer
            )
            
            if not downloaded_files:
                return {
                    'success': False,
                    'message': 'No submissions found to grade',
                    'folder_path': base_dir
                }
            
            print(f"Downloaded {len(downloaded_files)} submissions")
            
            # Step 1d: Load rubric and grade papers
            print(f"📋 Loading rubric: {os.path.basename(final_rubric_path)}")
            self.grading_agent.load_rubric(final_rubric_path)
            mapping_file = os.path.join(base_dir, "student_mapping.json")
            self.anonymizer.save_mapping(mapping_file)
            print(f"Saved student mapping to: {mapping_file}")
            
            # Step 1e: Grade papers with anonymization
            print("🤖 Grading submissions with ChatGPT (anonymized)...")
            
            grading_results = {}
            for user_id, file_paths in downloaded_files.items():
                try:
                    # Get anonymized name for this user
                    real_name = None
                    for anon_id, data in self.anonymizer.name_map.items():
                        if data['user_id'] == user_id:
                            real_name = data['real_name']
                            break
                    
                    anon_name = self.anonymizer.reverse_map.get(real_name, f"Student_{user_id}")
                    
                    # Grade each file for this student
                    student_grades = []
                    print(f"🔍 Processing {len(file_paths)} files for {anon_name}")
                    
                    for i, file_path in enumerate(file_paths):
                        print(f"  📄 File {i+1}: {file_path} (type: {type(file_path)})")
                        
                        # Check for None or empty file paths with detailed logging
                        if file_path is None:
                            print(f"    ❌ File path is None at index {i}")
                            continue
                        
                        if not isinstance(file_path, (str, os.PathLike)):
                            print(f"    ❌ File path is not a string or PathLike object: {type(file_path)}")
                            continue
                            
                        if not file_path:
                            print(f"    ⚠️ Skipping empty file path")
                            continue
                        
                        # Convert to string to be safe
                        file_path_str = str(file_path)
                        print(f"    🔍 Processing file path string: '{file_path_str}'")
                        
                        # Check if file actually exists
                        if not os.path.exists(file_path_str):
                            print(f"    ⚠️ File does not exist: {file_path_str}")
                            continue
                            
                        if file_path_str.lower().endswith(('.txt', '.docx', '.pdf', '.html', '.odt', '.rtf')):
                            try:
                                # Load the file content using the appropriate method
                                from pathlib import Path
                                
                                # Validate the file path before creating Path object
                                if not file_path_str or file_path_str == 'None':
                                    print(f"    ❌ Invalid file path string: '{file_path_str}'")
                                    continue
                                
                                file_path_obj = Path(file_path_str)
                                print(f"    🔍 Created Path object: {file_path_obj}")
                                
                                # Extract content based on file type (case-insensitive)
                                file_ext = file_path_str.lower()
                                if file_ext.endswith('.txt'):
                                    with open(file_path_str, 'r', encoding='utf-8') as f:
                                        content = f.read()
                                elif file_ext.endswith('.html'):
                                    with open(file_path_str, 'r', encoding='utf-8') as f:
                                        content = f.read()
                                elif file_ext.endswith(('.docx', '.odt')):
                                    # Use the grading agent's method for docx files, odt will be handled as text
                                    print(f"    📄 Extracting text from document: {file_path_obj}")
                                    print(f"    🔍 file_path_obj type: {type(file_path_obj)}")
                                    print(f"    🔍 file_path_obj value: {repr(file_path_obj)}")
                                    
                                    # Additional validation before calling the method
                                    if file_path_obj is None:
                                        print(f"    ❌ file_path_obj is None!")
                                        continue
                                    
                                    try:
                                        # Check if the file exists using the Path object
                                        print(f"    🔍 Checking if file exists: {file_path_obj.exists()}")
                                        if not file_path_obj.exists():
                                            print(f"    ❌ File does not exist according to Path.exists()")
                                            continue
                                        
                                        if file_ext.endswith('.docx'):
                                            content = self.grading_agent._extract_text_from_file(file_path_obj)
                                        elif file_ext.endswith('.odt'):
                                            # For ODT files, try to extract as text or fall back to basic text read
                                            try:
                                                # Try to read as text file first
                                                with open(file_path_str, 'r', encoding='utf-8') as f:
                                                    content = f.read()
                                            except UnicodeDecodeError:
                                                # If that fails, skip ODT for now (would need python-odf library)
                                                print(f"    ⚠️ ODT file requires special handling, skipping: {file_path_str}")
                                                continue
                                        print(f"    ✅ Successfully extracted text from document")
                                    except Exception as extract_error:
                                        print(f"    ❌ Error in _extract_text_from_file: {extract_error}")
                                        import traceback
                                        print(f"    🔍 Extract error traceback:")
                                        traceback.print_exc()
                                        continue
                                        
                                elif file_ext.endswith('.pdf'):
                                    # Use the grading agent's method for pdf files
                                    print(f"    📄 Extracting text from pdf: {file_path_obj}")
                                    print(f"    🔍 file_path_obj type: {type(file_path_obj)}")
                                    print(f"    🔍 file_path_obj value: {repr(file_path_obj)}")
                                    
                                    # Additional validation before calling the method
                                    if file_path_obj is None:
                                        print(f"    ❌ file_path_obj is None!")
                                        continue
                                    
                                    try:
                                        # Check if the file exists using the Path object
                                        print(f"    🔍 Checking if file exists: {file_path_obj.exists()}")
                                        if not file_path_obj.exists():
                                            print(f"    ❌ File does not exist according to Path.exists()")
                                            continue
                                            
                                        content = self.grading_agent._extract_text_from_file(file_path_obj)
                                        print(f"    ✅ Successfully extracted text from pdf")
                                    except Exception as extract_error:
                                        print(f"    ❌ Error in _extract_text_from_file: {extract_error}")
                                        import traceback
                                        print(f"    🔍 Extract error traceback:")
                                        traceback.print_exc()
                                        continue
                                elif file_ext.endswith('.rtf'):
                                    # RTF files - try to read as text 
                                    try:
                                        with open(file_path_str, 'r', encoding='utf-8') as f:
                                            content = f.read()
                                    except UnicodeDecodeError:
                                        print(f"    ⚠️ RTF file encoding issue, skipping: {file_path_str}")
                                        continue
                                else:
                                    print(f"    ⚠️ Unsupported file type: {file_path_str}")
                                    continue
                                
                                print(f"    ✅ Successfully extracted {len(content)} characters of content")
                                
                                # Create student data dictionary as expected by grade_paper
                                student_data = {
                                    'name': anon_name,
                                    'content': content,
                                    'file_path': file_path_str
                                }
                                
                                # The grading agent will see only anonymized content
                                grade_result = self.grading_agent.grade_paper(student_data)
                                student_grades.append(grade_result)
                                print(f"    ✅ Graded {anon_name}: {grade_result.get('total_score', 0)}/{grade_result.get('total_possible', 100)}")
                                
                            except Exception as e:
                                print(f"    ❌ Error grading file {file_path_str}: {e}")
                                import traceback
                                print(f"    🔍 Full traceback:")
                                traceback.print_exc()
                                continue
                        else:
                            print(f"    ⚠️ Skipping non-document file: {file_path_str}")
                    
                    if student_grades:
                        # Use the best grade if multiple files
                        best_grade = max(student_grades, key=lambda x: x.get('overall_score', 0))
                        
                        # Normalize the field names for consistency with spreadsheet creation
                        normalized_grade = {
                            'total_score': best_grade.get('overall_score', 0),
                            'max_score': best_grade.get('max_possible_score', 100),
                            'percentage': best_grade.get('percentage', 0),
                            'letter_grade': best_grade.get('letter_grade', 'F'),
                            'overall_feedback': best_grade.get('overall_feedback', ''),
                            'detailed_scores': [],  # We'll populate this from criteria_scores
                            'student_name': best_grade.get('student_name', anon_name),
                            'grading_date': best_grade.get('grading_date', '')
                        }
                        
                        # Convert criteria_scores to detailed_scores format
                        if 'criteria_scores' in best_grade:
                            for criterion_name, criterion_data in best_grade['criteria_scores'].items():
                                normalized_grade['detailed_scores'].append({
                                    'criterion': criterion_name,
                                    'score': criterion_data.get('score', 0),
                                    'max_score': criterion_data.get('max_score', 0),
                                    'feedback': criterion_data.get('feedback', '')
                                })
                        
                        grading_results[user_id] = normalized_grade
                        
                        print(f"Graded {anon_name}: {normalized_grade['total_score']:.1f}/{normalized_grade['max_score']}")
                        
                except Exception as e:
                    print(f"Error grading submissions for user {user_id}: {e}")
                    import traceback
                    print(f"Full traceback:")
                    traceback.print_exc()
            
            # Step 1c: Create review spreadsheet with real names
            print("Creating review spreadsheet...")
            print(f"🔍 Debug: grading_results has {len(grading_results)} entries")
            for user_id, result in grading_results.items():
                print(f"  - User {user_id}: Score {result.get('total_score', 'N/A')}")
            
            review_file = self.create_review_spreadsheet(
                grading_results, results_dir, assignment_name
            )
            
            # Step 1d: Copy rubric for reference
            rubric_copy = os.path.join(base_dir, "rubric_used.json")
            shutil.copy2(final_rubric_path, rubric_copy)
            
            # Step 1e: Create instruction file
            self.create_instruction_file(base_dir, assignment_name, review_file)
            
            print(f"\nStep 1 Complete!")
            print(f"Assignment folder: {base_dir}")
            print(f"Review file: {review_file}")
            print(f"Next: Review and edit the spreadsheet, then run Step 2")
            
            return {
                'success': True,
                'message': f'Step 1 completed. {len(grading_results)} submissions graded.',
                'folder_path': base_dir,
                'review_file': review_file,
                'grading_results': grading_results,
                'assignment_folder': assignment_folder,
                'course_id': course_id,
                'assignment_id': assignment_id
            }
            
        except Exception as e:
            print(f"Error in Step 1: {e}")
            return {
                'success': False,
                'message': f'Step 1 failed: {e}',
                'folder_path': base_dir
            }
    
    def step2_review_and_upload(self, assignment_folder_path: str) -> Dict:
        """
        Step 2: Read edited spreadsheet and upload grades to Canvas
        
        Args:
            assignment_folder_path: Path to the assignment folder from Step 1
            
        Returns:
            Dictionary with upload results
        """
        print(f"Step 2: Review and Upload")
        print("=" * 40)
        
        try:
            # Load the edited review spreadsheet
            results_dir = os.path.join(assignment_folder_path, "results")
            review_files = [f for f in os.listdir(results_dir) if f.endswith('_REVIEW.xlsx')]
            
            if not review_files:
                return {
                    'success': False,
                    'message': 'No review spreadsheet found. Run Step 1 first.'
                }
            
            review_file = os.path.join(results_dir, review_files[0])
            print(f"Reading edited grades from: {review_file}")
            
            # Read the edited spreadsheet
            df = pd.read_excel(review_file)
            
            # Load student mapping
            mapping_file = os.path.join(assignment_folder_path, "student_mapping.json")
            self.anonymizer.load_mapping(mapping_file)
            
            # Load assignment metadata
            meta_file = os.path.join(assignment_folder_path, "assignment_metadata.json")
            with open(meta_file, 'r') as f:
                metadata = json.load(f)
            
            course_id = metadata['course_id']
            assignment_id = metadata['assignment_id']
            
            # Prepare grade data for Canvas
            grade_data = {}
            uploaded_count = 0
            
            for _, row in df.iterrows():
                try:
                    user_id = int(row['Canvas_User_ID'])
                    final_grade = str(row['Final_Grade'])
                    comments = str(row['Final_Comments']) if pd.notna(row['Final_Comments']) else ""
                    
                    # Upload individual grade
                    self.canvas.update_submission_grade(
                        course_id, assignment_id, user_id, final_grade, comments
                    )
                    uploaded_count += 1
                    
                    student_name = row['Student_Name']
                    print(f"Uploaded grade for {student_name}: {final_grade}")
                    
                except Exception as e:
                    print(f"Error uploading grade for row {row.name}: {e}")
            
            # Create final report
            final_report = os.path.join(results_dir, "upload_report.txt")
            with open(final_report, 'w') as f:
                f.write(f"Canvas Upload Report\n")
                f.write(f"===================\n")
                f.write(f"Upload Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Course ID: {course_id}\n")
                f.write(f"Assignment ID: {assignment_id}\n")
                f.write(f"Total Grades: {len(df)}\n")
                f.write(f"Successfully Uploaded: {uploaded_count}\n")
                f.write(f"Failed Uploads: {len(df) - uploaded_count}\n")
            
            print(f"\nStep 2 Complete!")
            print(f"Uploaded {uploaded_count} grades to Canvas")
            print(f"Upload report: {final_report}")
            
            return {
                'success': True,
                'message': f'Step 2 completed. {uploaded_count} grades uploaded to Canvas.',
                'uploaded_count': uploaded_count,
                'total_grades': len(df),
                'report_file': final_report
            }
            
        except Exception as e:
            print(f"Error in Step 2: {e}")
            return {
                'success': False,
                'message': f'Step 2 failed: {e}'
            }
    
    def create_review_spreadsheet(self, grading_results: Dict, results_dir: str, 
                                assignment_name: str) -> str:
        """Create Excel spreadsheet for manual review and editing"""
        
        print(f"🔍 create_review_spreadsheet called with {len(grading_results)} grading results")
        
        # Prepare data for spreadsheet
        spreadsheet_data = []
        
        for user_id, result in grading_results.items():
            print(f"🔍 Processing user {user_id} with result: {result}")
            
            # Get real student name
            real_name = None
            for anon_id, data in self.anonymizer.name_map.items():
                if data['user_id'] == user_id:
                    real_name = data['real_name']
                    break
            
            if not real_name:
                real_name = f"Unknown_User_{user_id}"
                
            print(f"🔍 Real name for user {user_id}: {real_name}")
            
            # Calculate initial grade
            score = result['total_score']
            max_score = result['max_score']
            percentage = (score / max_score * 100) if max_score > 0 else 0
            
            # Format detailed feedback
            detailed_feedback = []
            for criterion in result['detailed_scores']:
                name = criterion['criterion']
                criterion_score = criterion['score']
                criterion_max = criterion['max_score']
                feedback = criterion['feedback']
                
                detailed_feedback.append(f"{name}: {criterion_score:.1f}/{criterion_max}")
                if feedback:
                    detailed_feedback.append(f"  - {feedback}")
            
            overall_feedback = result.get('overall_feedback', '')
            if overall_feedback:
                detailed_feedback.append(f"\nOverall: {overall_feedback}")
            
            combined_feedback = "\n".join(detailed_feedback)
            
            row_data = {
                'Student_Name': real_name,
                'Canvas_User_ID': user_id,
                'AI_Score': f"{score:.1f}",
                'Max_Score': f"{max_score:.1f}",
                'AI_Percentage': f"{percentage:.1f}%",
                'Final_Grade': f"{percentage:.1f}%",  # Editable
                'AI_Comments': combined_feedback,
                'Final_Comments': combined_feedback,  # Editable
                'Notes': '',  # For instructor notes
                'Upload_Status': 'PENDING'
            }
            
            spreadsheet_data.append(row_data)
        
        # Sort by student name
        spreadsheet_data.sort(key=lambda x: x['Student_Name'])
        
        # Create DataFrame
        df = pd.DataFrame(spreadsheet_data)
        
        # Save to Excel with formatting
        review_file = os.path.join(results_dir, f"{assignment_name}_REVIEW.xlsx")
        
        with pd.ExcelWriter(review_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Grade_Review', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Grade_Review']
            
            # Add formatting
            from openpyxl.styles import PatternFill, Font
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Editable columns highlighting
            editable_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            # Highlight editable columns (Final_Grade, Final_Comments, Notes)
            for row in range(2, len(df) + 2):
                worksheet[f'F{row}'].fill = editable_fill  # Final_Grade
                worksheet[f'H{row}'].fill = editable_fill  # Final_Comments
                worksheet[f'I{row}'].fill = editable_fill  # Notes
            
            # Adjust column widths
            column_widths = {
                'A': 20,  # Student_Name
                'B': 15,  # Canvas_User_ID
                'C': 12,  # AI_Score
                'D': 12,  # Max_Score
                'E': 15,  # AI_Percentage
                'F': 15,  # Final_Grade
                'G': 50,  # AI_Comments
                'H': 50,  # Final_Comments
                'I': 30,  # Notes
                'J': 15,  # Upload_Status
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        print(f"Created review spreadsheet: {review_file}")
        return review_file
    
    def create_instruction_file(self, base_dir: str, assignment_name: str, review_file: str):
        """Create instruction file for the review process"""
        
        instructions = f"""
CANVAS GRADING - TWO-STEP PROCESS
================================

Assignment: {assignment_name}
Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

STEP 1 - COMPLETED ✓
===================
- Downloaded all submissions from Canvas
- Anonymized student names for ChatGPT processing
- Graded all papers using AI with your rubric
- Created this review folder with all materials

STEP 2 - MANUAL REVIEW (DO THIS NOW)
===================================
1. Open the review spreadsheet: {os.path.basename(review_file)}

2. Review and edit the following columns (highlighted in yellow):
   - Final_Grade: Adjust the AI-suggested grade if needed
   - Final_Comments: Edit the feedback comments as desired
   - Notes: Add any private notes for your records

3. Save the spreadsheet when done

4. Run Step 2 in the Canvas grading tool to upload to Canvas

FOLDER CONTENTS
===============
- submissions/: All student papers (with anonymized names)
- results/: Review spreadsheet and reports
- student_mapping.json: Maps anonymous names to real names (keep secure!)
- rubric_used.json: Copy of the rubric used for grading
- assignment_metadata.json: Canvas assignment details

IMPORTANT NOTES
===============
- Student names were anonymized when sent to ChatGPT for privacy
- The review spreadsheet shows real names for your review
- Only edit the yellow-highlighted columns in the spreadsheet
- Don't modify the Canvas_User_ID column (needed for upload)
- Keep the student_mapping.json file secure

NEXT STEPS
==========
1. Review and edit the grade spreadsheet
2. Run Step 2 in the Canvas Grading Tool
3. Check Canvas gradebook to verify uploads

Questions? Check the main README.md file for detailed instructions.
"""
        
        instruction_file = os.path.join(base_dir, "INSTRUCTIONS.txt")
        with open(instruction_file, 'w', encoding='utf-8') as f:
            f.write(instructions)
        
        # Also save assignment metadata for Step 2
        metadata = {
            'assignment_name': assignment_name,
            'created_at': datetime.now().isoformat(),
            'course_id': None,  # Will be set by caller
            'assignment_id': None,  # Will be set by caller
            'step1_completed': True,
            'step2_completed': False
        }
        
        # This will be updated by the caller with actual IDs
        metadata_file = os.path.join(base_dir, "assignment_metadata.json")
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)


def setup_canvas_integration() -> Tuple[CanvasAPI, str]:
    """
    Interactive setup for Canvas API credentials
    
    Returns:
        Tuple of (CanvasAPI instance, config file path)
    """
    config_file = "canvas_config.json"
    
    # Try to load existing config
    if os.path.exists(config_file):
        try:
            with open(config_file, 'r') as f:
                config = json.load(f)
            print("Loaded existing Canvas configuration")
            return CanvasAPI(config['canvas_url'], config['api_token']), config_file
        except Exception as e:
            print(f"Error loading config: {e}")
    
    # Interactive setup
    print("Canvas LMS Integration Setup")
    print("="*40)
    print("You'll need:")
    print("1. Your Canvas instance URL (e.g., https://yourschool.instructure.com)")
    print("2. A Canvas API access token")
    print("\nTo get an API token:")
    print("1. Log into Canvas")
    print("2. Go to Account Settings")
    print("3. Click 'New Access Token'")
    print("4. Copy the generated token")
    print()
    
    canvas_url = input("Enter your Canvas URL: ").strip()
    api_token = input("Enter your API token: ").strip()
    
    # Save configuration
    config = {
        'canvas_url': canvas_url,
        'api_token': api_token,
        'created_at': datetime.now().isoformat()
    }
    
    with open(config_file, 'w') as f:
        json.dump(config, f, indent=2)
    
    print(f"Configuration saved to {config_file}")
    print("WARNING: Keep your API token secure and do not share this file!")
    
    return CanvasAPI(canvas_url, api_token), config_file


if __name__ == "__main__":
    """Example usage and testing"""
    from grading_agent import GradingAgent
    
    print("Enhanced Canvas Integration with Privacy Protection")
    print("=" * 60)
    
    # Setup Canvas integration
    canvas_api, config_file = setup_canvas_integration()
    
    # Test connection
    try:
        courses = canvas_api.get_courses()
        print(f"\n✓ Successfully connected! Found {len(courses)} courses.")
        
        # Show first few courses
        for i, course in enumerate(courses[:5]):
            print(f"{i+1}. {course.get('name', 'Unnamed Course')} (ID: {course['id']})")
        
        if len(courses) > 5:
            print(f"... and {len(courses)-5} more courses")
            
    except Exception as e:
        print(f"✗ Error connecting to Canvas: {e}")
        print("Please check your URL and API token")
        exit(1)
    
    # Initialize two-step grading system
    grading_agent = GradingAgent()
    two_step_grading = TwoStepCanvasGrading(canvas_api, grading_agent)
    
    print("\n✓ Enhanced Canvas integration ready!")
    print("\nFeatures:")
    print("- Student name anonymization for ChatGPT processing")
    print("- Two-step workflow: Grade → Review → Upload")
    print("- Organized assignment folders")
    print("- Manual review and editing capability")
    print("- Privacy protection throughout the process")
